import openpyxl
import MySQLdb
import sys

try:
    book = openpyxl.load_workbook("/home/pi/Documents/Weather.xlsx")
    
    sheet = book.active
    
    database = MySQLdb.connect(host= "127.0.0.1", user= "user", passwd= "Pa$$word12345", db= "csit216")
    
    cursor = database.cursor()

except MySQLdb.Error as e:

    print("Error %d: %s" % (e.args[0], e.args[1]))

    sys.exit(1)

insert = "INSERT INTO `weather` (`date`, `temperature`, `humidity`) VALUES (%s, %s, %s)"

count = 0

for x in range(2, sheet.max_row):
    date = sheet.cell(x,1).value
    temp = sheet.cell(x,2).value
    hum = sheet.cell(x,3).value
 
    cursor.execute(insert,(date, temp, hum))
    
    count+=1

cursor.close()

database.commit()

database.close()

print(count +" records added")